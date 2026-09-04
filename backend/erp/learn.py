"""Builds a customer's profile from files they already have.

The alias table in `schema.yaml` was made by a person at 四維 who knew that
`代工原料卷號` and `L/C NO.` are both the lot number. Every new customer needs
the same table for their own suppliers, and asking them to write one is asking
for the work that made this project necessary in the first place.

They do not have to. Two things they already own carry the answer:

* **their alias table** (`key.xlsx` at 四維) — columns are the ERP fields, the
  cells under each are the spellings suppliers use. That *is* a profile, and
  importing it needs no model at all: `columns_from_alias_table` below reads it
  straight across, defining the column set and the aliases in one go.

* **reports they have already done by hand** — a COA next to the filled-in
  import workbook for it. Uploading a PDF alone would not do: it carries the
  supplier's column names but not the judgement about what they map to. Pairing
  it with the answer recovers exactly that judgement, and `samples_for_draft`
  packs the pairs up for a model to generalise from.

Nothing here writes a profile. Both paths produce a *draft* that a human edits
and saves — a learned alias that nobody looked at is the same
unverified-rule problem, one step further from anyone noticing.
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any

logger = logging.getLogger("printlens.erp.learn")

# Guards on a workbook that arrives from a browser. Both are far above any real
# COA or alias table and far below anything that would hurt.
MAX_ROWS = 5_000
MAX_COLUMNS = 60


class SheetError(ValueError):
    """The workbook could not be read as a table."""


def _cell(value: Any) -> str:
    """Text, verbatim. Numbers must not gain a `.0` — these are lot numbers."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_sheet(data: bytes, sheet: str = "") -> tuple[list[str], list[list[str]]]:
    """`(headers, rows)` from a workbook, as text.

    The header is the first row with more than one non-empty cell — these
    workbooks routinely open with a merged title row, and taking row 1 blindly
    produces a table whose only column is the customer's company name.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise SheetError(f"讀不出這個 Excel 檔（{e}）") from e

    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
        headers: list[str] = []
        rows: list[list[str]] = []
        for raw in ws.iter_rows(values_only=True):
            cells = [_cell(v) for v in (raw or ())][:MAX_COLUMNS]
            if not headers:
                if sum(1 for c in cells if c) > 1:
                    headers = cells
                continue
            if any(cells):
                rows.append(cells)
            if len(rows) >= MAX_ROWS:
                break
    finally:
        wb.close()

    if not headers:
        raise SheetError("這個工作表找不到表頭列")
    return headers, rows


def sheet_names(data: bytes) -> list[str]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True)
    except Exception as e:
        raise SheetError(f"讀不出這個 Excel 檔（{e}）") from e
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def rows_as_dicts(headers: list[str], rows: list[list[str]]) -> list[dict]:
    """Rows keyed by their header, dropping unnamed columns."""
    out = []
    for cells in rows:
        row = {
            h: (cells[i] if i < len(cells) else "")
            for i, h in enumerate(headers)
            if h
        }
        if any(row.values()):
            out.append(row)
    return out


# ── key.xlsx → columns ───────────────────────────────────────────────────────
def _key_for(name: str, used: set[str]) -> str:
    """A stable ASCII key for a column whose name is usually Chinese.

    The name is what the ERP template matches on, so it is what has to be
    preserved; the key is only an internal handle, and `col_1` is a perfectly
    good handle. Where the name is already ASCII, use it — a profile whose keys
    read `supplier_lot` is far easier to hand-edit afterwards.
    """
    slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")[:40]
    if not slug or not slug[0].isalpha():
        slug = ""
    candidate = slug or f"col_{len(used) + 1}"
    n = 2
    while candidate in used:
        candidate = f"{slug or 'col'}_{n}"
        n += 1
    used.add(candidate)
    return candidate


def columns_from_alias_table(headers: list[str], rows: list[list[str]]) -> list[dict]:
    """Read an alias table (`key.xlsx`) straight into a profile's columns.

    Column = ERP field, cells beneath = the supplier spellings for it. No model
    involved: this is a transcription, and a transcription that a model
    paraphrased would be worse.

    `required` is left off every column on purpose. Something has to be
    required or `normalise_rows` keeps every blank row a model emits, so the
    save is refused until a human ticks the columns that make a row real —
    which is a decision about their ERP, not one to guess.
    """
    used: set[str] = set()
    columns = []
    for i, name in enumerate(headers):
        if not name:
            continue
        aliases = []
        for cells in rows:
            value = cells[i] if i < len(cells) else ""
            if value and value != name:
                aliases.append(value)
        columns.append(
            {
                "key": _key_for(name, used),
                "name": name,
                "required": False,
                "description": "",
                "aliases": list(dict.fromkeys(aliases)),
            }
        )
    if not columns:
        raise SheetError("表頭是空的，讀不到任何欄位")
    return columns


# ── (report, answer) pairs → prompt material ─────────────────────────────────
# A model reading many long COAs drifts: by the fifth document it is
# generalising from what it remembers rather than what is in front of it. Ten
# is comfortably inside where the pairs still all fit and get read.
MAX_SAMPLES = 10
MAX_SAMPLE_CHARS = 40_000


def samples_for_draft(samples: list[dict]) -> str:
    """Render `(markdown, the rows a human already produced)` pairs for a model.

    Answer first, then the document: the task is to explain an answer that is
    already known, not to produce one. A sample with no answer is still worth
    sending — it contributes the supplier's column names — but it is labelled,
    so nothing invented for it can be mistaken for evidence.
    """
    blocks = []
    for i, s in enumerate(samples[:MAX_SAMPLES], 1):
        md = (s.get("markdown") or "")[:MAX_SAMPLE_CHARS]
        expected = s.get("expected_rows") or []
        blocks.append(f"## 樣本 {i}：{s.get('filename', '?')}")
        if expected:
            headers = list(expected[0].keys())
            blocks.append("")
            blocks.append("**這份報告人工整理出來的正確答案**（表頭就是他們的 ERP 欄位）：")
            blocks.append("")
            blocks.append("| " + " | ".join(headers) + " |")
            blocks.append("|" + "|".join(["---"] * len(headers)) + "|")
            for row in expected[:200]:
                blocks.append(
                    "| " + " | ".join(str(row.get(h, "")).replace("|", "／") for h in headers) + " |"
                )
        else:
            blocks.append("")
            blocks.append("**這份沒有對照答案**——只能拿來看這家供應商的欄名寫法，"
                          "不要為它編造對應關係。")
        blocks.append("")
        blocks.append("報告原文：")
        blocks.append("")
        blocks.append(md)
        blocks.append("")
    return "\n".join(blocks)


def merge_columns(base: list[dict], learned: list[dict]) -> list[dict]:
    """Fold a draft's aliases into an existing column set.

    Matched on `key` first and display `name` second, because a model asked to
    describe an existing profile will reliably reproduce the visible names and
    only usually reproduce the keys. Columns are never dropped here and their
    order never changes — the order *is* the ERP import template — so the worst
    a bad draft can do is suggest aliases somebody then deletes.
    """
    by_key = {c["key"]: c for c in base}
    by_name = {c["name"]: c for c in base}
    out = [dict(c, aliases=list(c.get("aliases") or [])) for c in base]
    index = {c["key"]: i for i, c in enumerate(out)}

    for col in learned:
        key, name = str(col.get("key", "")), str(col.get("name", ""))
        target = by_key.get(key) or by_name.get(name)
        if target is None:
            # A column the customer has and the base profile does not. Keep it
            # — for a new customer this is most of the profile.
            out.append(
                {
                    "key": key or _key_for(name, set(index)),
                    "name": name or key,
                    "required": bool(col.get("required")),
                    "description": str(col.get("description") or ""),
                    "aliases": list(dict.fromkeys(col.get("aliases") or [])),
                }
            )
            continue
        i = index[target["key"]]
        merged = out[i]["aliases"] + [
            a for a in (col.get("aliases") or []) if isinstance(a, str)
        ]
        out[i]["aliases"] = list(dict.fromkeys(a.strip() for a in merged if a.strip()))
        if not out[i].get("description") and col.get("description"):
            out[i]["description"] = str(col["description"]).strip()
    return out
