"""Turns mapped rows into the files the ERP import actually eats.

Two shapes, because the existing manual process uses both:

* **per-file xlsx** — `Sheet1` with the 7 ERP columns plus a `context` sheet
  carrying the source markdown. This mirrors the workbooks already sitting in
  the customer's `OUTPUT/` folder byte-for-byte in structure, so whatever
  currently consumes those keeps working.
* **batch xlsx / csv** — every job in one table with a 來源檔案 column in
  front, for importing a whole day's inspection reports in one go.

Everything is written as text. Lot numbers like `24102102` and results like
`0.30` are identifiers and measurements, not quantities: letting Excel coerce
them to floats drops the leading zero, rewrites `1130220M0057` in scientific
notation, and produces an ERP row nobody can trace back to the PDF.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import schema, store

_HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
_HEADER_FONT = Font(bold=True)

# Excel forbids these in a sheet name and caps it at 31 chars.
_BAD_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    base = _BAD_SHEET_CHARS.sub("_", name).strip() or "sheet"
    base = base[:31]
    candidate, n = base, 1
    while candidate.casefold() in used:
        suffix = f"_{n}"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.casefold())
    return candidate


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, headers: list[str], max_width: int = 42) -> None:
    for i, h in enumerate(headers, 1):
        width = len(str(h)) + 2
        for row in ws.iter_rows(min_col=i, max_col=i, min_row=2):
            v = row[0].value
            if v:
                # CJK glyphs render about twice as wide as the width unit.
                width = max(width, min(max_width, len(str(v)) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width


def _append_rows(ws, rows: list[dict], keys: list[str], prefix: list[str] = ()) -> None:
    for row in rows:
        ws.append([*prefix, *[str(row.get(k, "") or "") for k in keys]])


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ── Per-file workbook ────────────────────────────────────────────────────────
def job_xlsx(job_id: str) -> tuple[bytes, str]:
    """One job → (xlsx bytes, filename). Same layout as the existing OUTPUT/*.xlsx."""
    meta = store.get_meta(job_id)
    rows = store.get_rows(job_id)
    profile = meta.get("profile_id") or schema.DEFAULT_PROFILE
    headers = schema.headers(profile)
    keys = schema.keys(profile)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_header(ws, headers)
    _append_rows(ws, rows, keys)
    ws.freeze_panes = "A2"
    _autosize(ws, headers)

    ctx = wb.create_sheet("context")
    ctx.column_dimensions["A"].width = 110
    for line in store.get_markdown(job_id).splitlines():
        ctx.append([line])

    buf = io.BytesIO()
    wb.save(buf)
    stem = re.sub(r"\.[A-Za-z0-9]+$", "", meta.get("filename") or job_id)
    return buf.getvalue(), f"{stem}.xlsx"


# ── Batch workbook ───────────────────────────────────────────────────────────
def _profile_of(job_ids: list[str]) -> str:
    """The profile this workbook is written under — the first job's.

    One workbook has one header row, so a batch spanning two customers cannot
    be one file. The odd job out is listed on 未匯入 rather than silently
    written under the wrong columns.
    """
    for job_id in job_ids:
        try:
            return store.get_meta(job_id).get("profile_id") or schema.DEFAULT_PROFILE
        except store.JobNotFound:
            continue
    return schema.DEFAULT_PROFILE


def batch_xlsx(job_ids: list[str], *, only_reviewed: bool = True) -> tuple[bytes, str]:
    """Many jobs → one workbook: a combined sheet plus one sheet per file.

    `only_reviewed` holds back anything a human has not signed off on. These
    rows decide whether incoming material is accepted, so an unlooked-at
    mapping must not reach ERP by default — but it is listed on the 未匯入
    sheet, never silently dropped.
    """
    profile = _profile_of(job_ids)
    headers = schema.headers(profile)
    keys = schema.keys(profile)

    wb = Workbook()
    summary = wb.active
    summary.title = "彙總"
    _write_header(summary, ["來源檔案", *headers])

    used: set[str] = {"彙總"}
    skipped: list[tuple[str, str]] = []

    for job_id in job_ids:
        try:
            meta = store.get_meta(job_id)
        except store.JobNotFound:
            skipped.append((job_id, "找不到此 job"))
            continue
        rows = store.get_rows(job_id)
        name = meta.get("filename") or job_id
        if meta.get("kind", store.KIND_REPORT) != store.KIND_REPORT:
            skipped.append((name, "這是設定檔的學習樣本，不匯出"))
            continue
        if (meta.get("profile_id") or schema.DEFAULT_PROFILE) != profile:
            skipped.append((name, f"設定檔不同（{meta.get('profile_id')}），請分開匯出"))
            continue
        if not rows:
            skipped.append((name, meta.get("error") or f"狀態：{meta.get('status')}"))
            continue
        if only_reviewed and not meta.get("reviewed_at"):
            skipped.append((name, "尚未確認"))
            continue

        _append_rows(summary, rows, keys, prefix=[name])

        ws = wb.create_sheet(_safe_sheet_name(name, used))
        _write_header(ws, headers)
        _append_rows(ws, rows, keys)
        ws.freeze_panes = "A2"
        _autosize(ws, headers)

    summary.freeze_panes = "B2"
    _autosize(summary, ["來源檔案", *headers])

    # A file that produced no rows must not vanish silently — someone has to
    # know which PDFs still need a human.
    if skipped:
        ws = wb.create_sheet("未匯入")
        _write_header(ws, ["檔案", "原因"])
        for name, reason in skipped:
            ws.append([name, reason])
        _autosize(ws, ["檔案", "原因"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"ERP匯入_{_timestamp()}.xlsx"


# ── CSV ──────────────────────────────────────────────────────────────────────
def batch_csv(job_ids: list[str], *, only_reviewed: bool = True) -> tuple[bytes, str]:
    """Many jobs → one CSV, UTF-8 **with BOM**.

    Excel on a zh-TW Windows box opens a BOM-less UTF-8 CSV as cp950 and turns
    every Chinese column name into mojibake, which is how these files get
    reported as "broken export".

    CSV has nowhere to put a 未匯入 sheet, so an unreviewed job is simply
    absent here. The UI states the count before offering the link.
    """
    profile = _profile_of(job_ids)
    keys = schema.keys(profile)
    buf = io.StringIO(newline="")
    w = csv.writer(buf)
    w.writerow(["來源檔案", *schema.headers(profile)])
    for job_id in job_ids:
        try:
            meta = store.get_meta(job_id)
        except store.JobNotFound:
            continue
        if meta.get("kind", store.KIND_REPORT) != store.KIND_REPORT:
            continue
        if (meta.get("profile_id") or schema.DEFAULT_PROFILE) != profile:
            continue
        if only_reviewed and not meta.get("reviewed_at"):
            continue
        name = meta.get("filename") or job_id
        for row in store.get_rows(job_id):
            w.writerow([name, *[str(row.get(k, "") or "") for k in keys]])
    return buf.getvalue().encode("utf-8-sig"), f"ERP匯入_{_timestamp()}.csv"
