"""
Tests for the ERP import mode.

The normalisation tests are the important ones: rows arrive from an LLM, so
the shapes they come back in are the thing most likely to change without
warning, and a lot number silently turned into a float is an ERP row nobody
can trace back to the PDF.

    python -m pytest backend/tests/test_erp.py -q
"""

from __future__ import annotations

import io
import json
import sys
from pathlib import Path

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

BACKEND = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND))


@pytest.fixture()
def client(tmp_path, monkeypatch):
    """A client over a throwaway job store.

    ERP_JOBS_DIR is read at import time, so the module has to be imported
    fresh under the patched environment rather than reused across tests.
    """
    monkeypatch.setenv("ERP_JOBS_DIR", str(tmp_path / "jobs"))
    for mod in [m for m in sys.modules if m == "erp" or m.startswith("erp.")]:
        del sys.modules[mod]
    import erp

    app = FastAPI()
    app.include_router(erp.router)
    return TestClient(app)


def stage(client, filename="報告.pdf", markdown="# 檢驗報告\n", **kw):
    r = client.post(
        "/api/erp/jobs",
        json={"documents": [{"filename": filename, "markdown": markdown, **kw}]},
    )
    assert r.status_code == 201, r.text
    return r.json()["jobs"][0]["job_id"]


def review(client, job_id):
    """Sign a job off, the way a reviewer does before exporting."""
    r = client.post(f"/api/erp/jobs/{job_id}/review")
    assert r.status_code == 200, r.text
    return r.json()


def tiny_pdf(pages: int = 1) -> bytes:
    """A minimal but real PDF, built with the binding the app already uses."""
    import pypdfium2 as pdfium

    doc = pdfium.PdfDocument.new()
    for _ in range(pages):
        doc.new_page(200, 300)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


# ── Schema ───────────────────────────────────────────────────────────────────
def test_schema_is_the_seven_erp_columns_in_order(client):
    d = client.get("/api/erp/schema").json()
    assert [c["name"] for c in d["columns"]] == [
        "供應商批號",
        "檢驗項目",
        "單位",
        "規格",
        "規格上限",
        "規格下限",
        "檢驗結果",
    ]


def test_schema_carries_the_supplier_aliases(client):
    d = client.get("/api/erp/schema").json()
    by_key = {c["key"]: c for c in d["columns"]}
    # The two spellings that broke the old regex extractor most often.
    assert "L/C NO." in by_key["supplier_lot"]["aliases"]
    assert "实际值" in by_key["result"]["aliases"]
    # 簡體 reports are a normal case, not an exception.
    assert "批号" in by_key["supplier_lot"]["aliases"]


def test_schema_markdown_renders_for_the_mcp_resource(client):
    body = client.get("/api/erp/schema.md").text
    assert "ERP 匯入欄位定義" in body
    assert "`supplier_lot`" in body


# ── Staging ──────────────────────────────────────────────────────────────────
def test_staged_job_starts_pending_and_keeps_its_markdown(client):
    job_id = stage(client, markdown="# 南寶\n| 批號 | 24102102 |")
    d = client.get(f"/api/erp/jobs/{job_id}").json()
    assert d["status"] == "pending"
    assert "24102102" in d["markdown"]
    assert d["rows"] == []


def test_empty_markdown_is_recorded_as_failed_not_dropped(client):
    """A file OCR could not read still needs to be visible to a human."""
    job_id = stage(client, filename="掃描件.pdf", markdown="   ")
    d = client.get(f"/api/erp/jobs/{job_id}").json()
    assert d["status"] == "failed"
    assert d["error"]


def test_jobs_are_listed_newest_first_and_filter_by_status(client):
    stage(client, filename="a.pdf")
    b = stage(client, filename="b.pdf", markdown="")
    assert client.get("/api/erp/jobs").json()["count"] == 2
    failed = client.get("/api/erp/jobs", params={"status": "failed"}).json()
    assert [j["job_id"] for j in failed["jobs"]] == [b]


def test_batch_id_groups_an_upload(client):
    client.post(
        "/api/erp/jobs",
        json={
            "batch_id": "batch-1",
            "documents": [
                {"filename": "a.pdf", "markdown": "x"},
                {"filename": "b.pdf", "markdown": "y"},
            ],
        },
    )
    stage(client, filename="other.pdf")
    d = client.get("/api/erp/jobs", params={"batch_id": "batch-1"}).json()
    assert d["count"] == 2


# ── Dual-engine output ───────────────────────────────────────────────────────
def test_both_engine_outputs_are_kept_when_dual_mode_ran(client):
    """Same document, two renderings — the second is the digit cross-check."""
    job_id = stage(
        client,
        markdown="| 批號 | 2410Z1O2 |",
        engine="dual",
        alt_markdown="| 批號 | 24102102 |",
        alt_engine="fastdoc",
    )
    d = client.get(f"/api/erp/jobs/{job_id}").json()
    assert d["has_alt"] is True
    assert d["alt_engine"] == "fastdoc"
    assert "2410Z1O2" in d["markdown"]
    assert "24102102" in d["alt_markdown"]


def test_alt_variant_is_served_separately(client):
    job_id = stage(client, markdown="marker 版", alt_markdown="fastdoc 版")
    assert client.get(f"/api/erp/jobs/{job_id}/markdown").text == "marker 版"
    r = client.get(f"/api/erp/jobs/{job_id}/markdown", params={"variant": "alt"})
    assert r.text == "fastdoc 版"


def test_single_engine_jobs_report_no_alt(client):
    """Most reports are scans with no text layer — there is nothing to compare."""
    job_id = stage(client, markdown="只有一種輸出")
    d = client.get(f"/api/erp/jobs/{job_id}").json()
    assert d["has_alt"] is False
    assert d["alt_markdown"] == ""
    assert d["alt_engine"] == ""


def test_blank_alt_markdown_does_not_count_as_dual(client):
    job_id = stage(client, markdown="x", alt_markdown="   ", alt_engine="fastdoc")
    d = client.get(f"/api/erp/jobs/{job_id}").json()
    assert d["has_alt"] is False
    assert d["alt_engine"] == ""


def test_oversized_alt_markdown_is_refused(client):
    from erp import routes

    r = client.post(
        "/api/erp/jobs",
        json={
            "documents": [
                {
                    "filename": "big.pdf",
                    "markdown": "ok",
                    "alt_markdown": "x" * (routes.MAX_MARKDOWN_CHARS + 1),
                }
            ]
        },
    )
    assert r.status_code == 413


# ── Row normalisation ────────────────────────────────────────────────────────
def test_rows_accept_both_english_keys_and_chinese_column_names(client):
    job_id = stage(client)
    r = client.put(
        f"/api/erp/jobs/{job_id}/rows",
        json={
            "rows": [
                {"test_item": "固成份", "result": "40.12"},
                {"檢驗項目": "黏度", "檢驗結果": "14550"},
            ]
        },
    )
    rows = r.json()["rows"]
    assert [x["test_item"] for x in rows] == ["固成份", "黏度"]
    assert [x["result"] for x in rows] == ["40.12", "14550"]


def test_numeric_values_are_stringified_so_excel_cannot_coerce_them(client):
    """A lot number is an identifier. 24102102 must not become 24102102.0."""
    job_id = stage(client)
    r = client.put(
        f"/api/erp/jobs/{job_id}/rows",
        json={"rows": [{"supplier_lot": 24102102, "test_item": "x", "result": 40.12}]},
    )
    row = r.json()["rows"][0]
    assert row["supplier_lot"] == "24102102"
    assert row["result"] == "40.12"


def test_rows_with_no_required_field_are_dropped_and_reported(client):
    job_id = stage(client)
    r = client.put(
        f"/api/erp/jobs/{job_id}/rows",
        json={"rows": [{"unit": "%"}, {"test_item": "固成份", "result": "40"}]},
    )
    d = r.json()
    assert d["row_count"] == 1
    assert any("必填" in w for w in d["warnings"])


def test_unknown_keys_are_ignored_and_reported(client):
    job_id = stage(client)
    d = client.put(
        f"/api/erp/jobs/{job_id}/rows",
        json={"rows": [{"test_item": "x", "result": "1", "confidence": 0.9}]},
    ).json()
    assert "confidence" not in d["rows"][0]
    assert any("confidence" in w for w in d["warnings"])


def test_a_row_with_a_result_but_no_spec_survives(client):
    """The single most-reported miss of the old extractor: 規格 blank, 結果 present."""
    job_id = stage(client)
    d = client.put(
        f"/api/erp/jobs/{job_id}/rows",
        json={"rows": [{"test_item": "外觀", "spec": "", "result": "淡黃色透明液體"}]},
    ).json()
    assert d["row_count"] == 1
    assert d["rows"][0]["result"] == "淡黃色透明液體"


def test_resubmitting_rows_replaces_rather_than_appends(client):
    job_id = stage(client)
    body = {"rows": [{"test_item": "a", "result": "1"}, {"test_item": "b", "result": "2"}]}
    client.put(f"/api/erp/jobs/{job_id}/rows", json=body)
    d = client.put(
        f"/api/erp/jobs/{job_id}/rows", json={"rows": [{"test_item": "a", "result": "9"}]}
    ).json()
    assert d["row_count"] == 1
    assert d["status"] == "mapped"


# ── Export ───────────────────────────────────────────────────────────────────
def test_job_export_has_sheet1_and_context(client):
    import io

    from openpyxl import load_workbook

    job_id = stage(client, markdown="原文第一行\n原文第二行")
    client.put(
        f"/api/erp/jobs/{job_id}/rows",
        json={"rows": [{"supplier_lot": "24102102", "test_item": "固成份", "result": "40.12"}]},
    )
    r = client.get(f"/api/erp/jobs/{job_id}/export.xlsx")
    assert r.status_code == 200

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Sheet1", "context"]
    assert wb["Sheet1"]["A1"].value == "供應商批號"
    assert wb["Sheet1"]["A2"].value == "24102102"
    assert wb["context"]["A1"].value == "原文第一行"


def test_batch_export_lists_unmapped_files_rather_than_dropping_them(client):
    from openpyxl import load_workbook

    good = stage(client, filename="good.pdf")
    bad = stage(client, filename="bad.pdf", markdown="")
    client.put(
        f"/api/erp/jobs/{good}/rows", json={"rows": [{"test_item": "x", "result": "1"}]}
    )
    review(client, good)
    r = client.get("/api/erp/export.xlsx", params={"job_ids": f"{good},{bad}"})
    wb = load_workbook(io.BytesIO(r.content))
    assert "彙總" in wb.sheetnames
    assert "未匯入" in wb.sheetnames
    assert wb["未匯入"]["A2"].value == "bad.pdf"


def test_csv_export_is_utf8_with_bom_for_excel(client):
    """Without the BOM, Excel on zh-TW Windows opens it as cp950 → mojibake."""
    job_id = stage(client)
    client.put(f"/api/erp/jobs/{job_id}/rows", json={"rows": [{"test_item": "外觀", "result": "OK"}]})
    review(client, job_id)
    r = client.get("/api/erp/export.csv", params={"job_ids": job_id})
    assert r.content.startswith(b"\xef\xbb\xbf")
    assert "檢驗項目" in r.content.decode("utf-8-sig")


# ── Source PDF & page images ─────────────────────────────────────────────────
def test_source_pdf_is_stored_and_its_pages_render(client):
    job_id = stage(client)
    r = client.post(
        f"/api/erp/jobs/{job_id}/source",
        files={"file": ("報告.pdf", tiny_pdf(2), "application/pdf")},
    )
    assert r.status_code == 201, r.text
    assert r.json()["has_source"] is True
    assert r.json()["page_count"] == 2

    r = client.get(f"/api/erp/jobs/{job_id}/page/1.png")
    assert r.status_code == 200
    assert r.headers["content-type"] == "image/png"
    assert r.content.startswith(b"\x89PNG")


def test_page_beyond_the_end_is_404_rather_than_a_blank_image(client):
    job_id = stage(client)
    client.post(
        f"/api/erp/jobs/{job_id}/source",
        files={"file": ("報告.pdf", tiny_pdf(1), "application/pdf")},
    )
    assert client.get(f"/api/erp/jobs/{job_id}/page/5.png").status_code == 404


def test_a_job_with_no_pdf_serves_no_page(client):
    """The review pane falls back to the markdown; it must not 500 here."""
    job_id = stage(client)
    assert client.get(f"/api/erp/jobs/{job_id}/page/1.png").status_code == 404


def test_non_pdf_upload_is_refused_on_content_not_extension(client):
    job_id = stage(client)
    r = client.post(
        f"/api/erp/jobs/{job_id}/source",
        files={"file": ("報告.pdf", b"MZ\x90\x00 this is an exe", "application/pdf")},
    )
    assert r.status_code == 400


def test_stored_pdfs_are_pruned_once_the_store_outgrows_its_disk_budget(client):
    """Counting jobs stopped bounding disk the moment PDFs were kept too."""
    from erp import store

    old_budget = store.MAX_TOTAL_BYTES
    store.MAX_TOTAL_BYTES = 100_000
    try:
        first = stage(client, filename="oldest.pdf")
        client.post(
            f"/api/erp/jobs/{first}/source",
            files={"file": ("oldest.pdf", tiny_pdf(1) + b"%" + b"0" * 250_000, "application/pdf")},
        )
        # Staging the next document is what runs the prune.
        stage(client, filename="newest.pdf")
        assert client.get(f"/api/erp/jobs/{first}").status_code == 404
    finally:
        store.MAX_TOTAL_BYTES = old_budget


def test_oversized_pdf_is_refused(client):
    from erp import store

    job_id = stage(client)
    r = client.post(
        f"/api/erp/jobs/{job_id}/source",
        files={
            "file": (
                "big.pdf",
                b"%PDF-1.7\n" + b"0" * store.SOURCE_MAX_BYTES,
                "application/pdf",
            )
        },
    )
    assert r.status_code == 413


# ── Review sign-off ──────────────────────────────────────────────────────────
def test_unreviewed_rows_are_held_back_from_the_batch_export(client):
    """These rows decide whether material is accepted — nobody's word for it."""
    from openpyxl import load_workbook

    job_id = stage(client, filename="待確認.pdf")
    client.put(
        f"/api/erp/jobs/{job_id}/rows", json={"rows": [{"test_item": "x", "result": "1"}]}
    )

    r = client.get("/api/erp/export.xlsx", params={"job_ids": job_id})
    wb = load_workbook(io.BytesIO(r.content))
    assert wb["彙總"]["A2"].value is None
    assert wb["未匯入"]["B2"].value == "尚未確認"

    review(client, job_id)
    r = client.get("/api/erp/export.xlsx", params={"job_ids": job_id})
    wb = load_workbook(io.BytesIO(r.content))
    assert wb["彙總"]["A2"].value == "待確認.pdf"


def test_only_reviewed_false_exports_everything(client):
    """知識通's preview link needs the pre-sign-off view."""
    from openpyxl import load_workbook

    job_id = stage(client)
    client.put(
        f"/api/erp/jobs/{job_id}/rows", json={"rows": [{"test_item": "x", "result": "1"}]}
    )
    r = client.get(
        "/api/erp/export.xlsx", params={"job_ids": job_id, "only_reviewed": "false"}
    )
    wb = load_workbook(io.BytesIO(r.content))
    assert wb["彙總"]["A2"].value == "報告.pdf"


def test_editing_rows_after_signing_off_clears_the_sign_off(client):
    """Whoever confirmed did not see the new rows."""
    job_id = stage(client)
    rows = {"rows": [{"test_item": "x", "result": "1"}]}
    client.put(f"/api/erp/jobs/{job_id}/rows", json=rows)
    review(client, job_id)
    assert client.get(f"/api/erp/jobs/{job_id}").json()["reviewed_at"]

    client.put(f"/api/erp/jobs/{job_id}/rows", json={"rows": [{"test_item": "y", "result": "2"}]})
    assert client.get(f"/api/erp/jobs/{job_id}").json()["reviewed_at"] is None


def test_sign_off_can_be_taken_back(client):
    job_id = stage(client)
    client.put(f"/api/erp/jobs/{job_id}/rows", json={"rows": [{"test_item": "x", "result": "1"}]})
    review(client, job_id)
    assert client.delete(f"/api/erp/jobs/{job_id}/review").status_code == 200
    assert client.get(f"/api/erp/jobs/{job_id}").json()["reviewed_at"] is None


# ── Backend-driven mapping (local / company LLM) ─────────────────────────────
@pytest.fixture()
def fake_llm(monkeypatch):
    """Swap the HTTP call out for a canned reply.

    The point of these tests is the plumbing around the model — normalisation,
    the notes, what happens on failure — not the model, so nothing here should
    need Ollama running.
    """
    from erp import llm

    calls = []

    def install(reply, provider="ollama"):
        async def runner(model, system, user, fmt=None):
            calls.append({"model": model, "system": system, "user": user, "fmt": fmt})
            if isinstance(reply, Exception):
                raise reply
            # Adapters return the model's raw text; llm._parse does the reading.
            return reply if isinstance(reply, str) else json.dumps(reply)

        monkeypatch.setattr(llm, "PROVIDER_ORDER", [provider])
        monkeypatch.setitem(llm._RUNNERS, provider, runner)
        return calls

    return install


def test_llm_mapping_stores_normalised_rows(client, fake_llm):
    fake_llm(
        {
            "rows": [{"檢驗項目": "固成份", "檢驗結果": "40.12"}],
            "notes": "第 3 列看不清楚",
        }
    )
    job_id = stage(client, markdown="| 項目 | 結果 |\n| 固成份 | 40.12 |")

    r = client.post(f"/api/erp/jobs/{job_id}/map", json={})
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["status"] == "mapped"
    # Chinese header in, English key out — and a float coerced to text, or the
    # lot number beside it would reach Excel as a number.
    assert d["rows"] == [
        {
            "supplier_lot": "",
            "test_item": "固成份",
            "unit": "",
            "spec": "",
            "spec_max": "",
            "spec_min": "",
            "result": "40.12",
        }
    ]
    assert "第 3 列看不清楚" in d["notes"]
    assert d["mapped_by"].startswith("ollama／")


def test_the_prompt_carries_the_schema_and_the_known_traps(client, fake_llm):
    calls = fake_llm({"rows": [{"test_item": "x", "result": "1"}]})
    job_id = stage(client, markdown="# 報告")
    client.post(f"/api/erp/jobs/{job_id}/map", json={})

    prompt = calls[0]["user"]
    assert "L/C NO." in prompt          # the alias list
    assert "版型與已知陷阱" in prompt    # report-patterns.md
    assert "不要編造" in calls[0]["system"]


def test_a_failed_run_leaves_the_job_for_知識通(client, fake_llm):
    """A model that cannot be reached is not a broken document."""
    fake_llm(RuntimeError("connection refused"))
    job_id = stage(client)

    r = client.post(f"/api/erp/jobs/{job_id}/map", json={})
    assert r.status_code == 502
    meta = client.get(f"/api/erp/jobs/{job_id}").json()
    assert meta["status"] == "pending"
    assert meta["mapping_state"] == "error"
    assert "connection refused" in meta["mapping_error"]


def test_mapping_with_no_provider_configured_says_so(client, monkeypatch):
    from erp import llm

    monkeypatch.setattr(llm, "PROVIDER_ORDER", [])
    job_id = stage(client)
    r = client.post(f"/api/erp/jobs/{job_id}/map", json={})
    assert r.status_code == 502
    assert "知識通" in r.json()["detail"]


def test_batch_mapping_covers_the_pending_queue(client, fake_llm):
    fake_llm({"rows": [{"test_item": "外觀", "result": "合格"}]})
    a = stage(client, filename="a.pdf")
    b = stage(client, filename="b.pdf")

    r = client.post("/api/erp/map", json={})
    assert r.status_code == 202
    assert sorted(r.json()["job_ids"]) == sorted([a, b])
    # TestClient runs background tasks before returning, so both are done.
    for job_id in (a, b):
        assert client.get(f"/api/erp/jobs/{job_id}").json()["status"] == "mapped"


def test_a_down_server_still_lists_models(client, monkeypatch):
    """An empty dropdown reads as 'unsupported'; the curated list reads as 'offline'."""
    from erp import llm

    llm._catalog.clear()
    monkeypatch.setattr(
        llm.httpx, "get", lambda *a, **kw: (_ for _ in ()).throw(OSError("no route"))
    )
    d = client.get("/api/erp/llm").json()
    assert d["providers"]["ollama"]["models"] == llm.PROVIDER_MODELS["ollama"]
    assert d["providers"]["ollama"]["error"]


# ── Customer profiles ────────────────────────────────────────────────────────
def xlsx(rows: list[list], title: str = "Sheet1") -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = title
    for row in rows:
        wb.active.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_profile(client, profile, columns, name="測試客戶"):
    return client.put(
        f"/api/erp/profiles/{profile}",
        json={"name": name, "columns": columns, "rules": []},
    )


def test_alias_table_import_reads_the_customers_own_key_xlsx(client):
    """Their key.xlsx *is* a profile — columns are the fields, cells the aliases."""
    book = xlsx(
        [
            ["爐號", "項目", "實測值"],
            ["Lot No.", "Test Item", "Result"],
            ["批号", "檢驗項目", "实际值"],
            ["", "Item", ""],
        ]
    )
    r = client.post(
        "/api/erp/profiles/acme/alias-table",
        files={"file": ("key.xlsx", book, "application/vnd.ms-excel")},
    )
    assert r.status_code == 200, r.text
    cols = r.json()["draft"]["columns"]
    assert [c["name"] for c in cols] == ["爐號", "項目", "實測值"]
    assert cols[0]["aliases"] == ["Lot No.", "批号"]
    # A blank cell is not an alias, and the tail of a short column is not either.
    assert cols[2]["aliases"] == ["Result", "实际值"]


def test_a_profile_with_no_required_column_is_refused(client):
    """Without one, normalise_rows keeps every blank row a model emits."""
    r = save_profile(
        client, "acme", [{"key": "a", "name": "甲", "aliases": []}]
    )
    assert r.status_code == 400
    assert "必填" in r.json()["detail"]


def test_the_default_profile_cannot_be_overwritten_through_the_api(client):
    r = save_profile(client, "default", [{"key": "a", "name": "甲", "required": True}])
    assert r.status_code == 400


def test_a_customers_columns_become_the_export_headers(client, tmp_path, monkeypatch):
    """Profiles own the column set, not just the aliases — templates differ."""
    from openpyxl import load_workbook

    assert save_profile(
        client,
        "acme",
        [
            {"key": "heat_no", "name": "爐號", "required": True, "aliases": ["Heat No."]},
            {"key": "value", "name": "實測值", "required": True, "aliases": []},
        ],
    ).status_code == 200

    r = client.post(
        "/api/erp/jobs",
        json={"profile_id": "acme", "documents": [{"filename": "a.pdf", "markdown": "x"}]},
    )
    job_id = r.json()["jobs"][0]["job_id"]
    client.put(
        f"/api/erp/jobs/{job_id}/rows",
        json={"rows": [{"爐號": "H-1", "實測值": "40.12", "test_item": "不在這個設定檔裡"}]},
    )
    review(client, job_id)

    wb = load_workbook(io.BytesIO(client.get("/api/erp/export.xlsx", params={"job_ids": job_id}).content))
    assert [c.value for c in wb["彙總"][1]] == ["來源檔案", "爐號", "實測值"]
    assert wb["彙總"]["B2"].value == "H-1"


def test_a_batch_spanning_two_profiles_does_not_silently_mix_them(client):
    """One workbook has one header row; the odd one out is listed, not written."""
    from openpyxl import load_workbook

    save_profile(client, "acme", [{"key": "heat_no", "name": "爐號", "required": True}])
    a = stage(client, filename="default.pdf")
    client.put(f"/api/erp/jobs/{a}/rows", json={"rows": [{"test_item": "x", "result": "1"}]})
    review(client, a)

    r = client.post(
        "/api/erp/jobs",
        json={"profile_id": "acme", "documents": [{"filename": "acme.pdf", "markdown": "x"}]},
    )
    b = r.json()["jobs"][0]["job_id"]
    client.put(f"/api/erp/jobs/{b}/rows", json={"rows": [{"爐號": "H-1"}]})
    review(client, b)

    wb = load_workbook(io.BytesIO(client.get("/api/erp/export.xlsx", params={"job_ids": f"{a},{b}"}).content))
    assert wb["彙總"]["A2"].value == "default.pdf"
    assert wb["未匯入"]["A2"].value == "acme.pdf"
    assert "設定檔不同" in wb["未匯入"]["B2"].value


def test_training_samples_stay_out_of_the_review_queue_and_the_export(client):
    """A sample is a document with a known answer, not work waiting for anyone."""
    from openpyxl import load_workbook

    r = client.post(
        "/api/erp/profiles/acme/samples",
        json={"documents": [{"filename": "sample.pdf", "markdown": "# 樣本"}]},
    )
    assert r.status_code == 201
    sample_id = r.json()["samples"][0]["job_id"]

    assert client.get("/api/erp/jobs").json()["count"] == 0
    assert client.get("/api/erp/profiles/acme/samples").json()["samples"][0]["job_id"] == sample_id

    client.put(f"/api/erp/jobs/{sample_id}/rows", json={"rows": [{"test_item": "x", "result": "1"}]})
    review(client, sample_id)
    wb = load_workbook(io.BytesIO(client.get("/api/erp/export.xlsx", params={"job_ids": sample_id}).content))
    assert "學習樣本" in wb["未匯入"]["B2"].value


def test_the_answer_workbook_is_read_verbatim(client):
    """This is the half a PDF cannot supply; interpreting it is the job being learned."""
    r = client.post(
        "/api/erp/profiles/acme/samples",
        json={"documents": [{"filename": "s.pdf", "markdown": "# 樣本"}]},
    )
    sample_id = r.json()["samples"][0]["job_id"]

    book = xlsx([["供應商批號", "檢驗項目", "檢驗結果"], [24102102, "固成份", 40.12]])
    r = client.post(
        f"/api/erp/jobs/{sample_id}/expected",
        files={"file": ("done.xlsx", book, "application/vnd.ms-excel")},
    )
    assert r.status_code == 200, r.text
    # A lot number that came back as 24102102.0 would be a different lot number.
    assert r.json()["expected_rows"] == [
        {"供應商批號": "24102102", "檢驗項目": "固成份", "檢驗結果": "40.12"}
    ]
    assert r.json()["expected_row_count"] == 1


def test_drafting_without_a_single_answer_says_what_is_missing(client, fake_llm):
    """PDFs alone carry column names but no judgement about what they map to."""
    fake_llm({"columns": []})
    client.post(
        "/api/erp/profiles/acme/samples",
        json={"documents": [{"filename": "s.pdf", "markdown": "# 樣本"}]},
    )
    r = client.post("/api/erp/profiles/acme/draft", json={})
    assert r.status_code == 400
    assert "對照答案" in r.json()["detail"]


def test_a_draft_can_add_aliases_but_never_drop_a_column(client, fake_llm):
    fake_llm(
        {
            "columns": [
                {"key": "supplier_lot", "name": "供應商批號", "aliases": ["爐號"]},
                {"key": "heat_treat", "name": "熱處理", "aliases": ["Heat Treatment"]},
            ],
            "notes": "第 2 份對不起來",
        }
    )
    r = client.post(
        "/api/erp/profiles/default/samples",
        json={"documents": [{"filename": "s.pdf", "markdown": "# 樣本"}]},
    )
    sample_id = r.json()["samples"][0]["job_id"]
    assert (
        client.post(
            f"/api/erp/jobs/{sample_id}/expected",
            files={
                "file": (
                    "done.xlsx",
                    xlsx([["供應商批號", "檢驗結果"], ["A1", "40.12"]]),
                    "application/vnd.ms-excel",
                )
            },
        ).status_code
        == 200
    )

    r = client.post("/api/erp/profiles/default/draft", json={})
    assert r.status_code == 200, r.text
    cols = {c["name"]: c for c in r.json()["draft"]["columns"]}
    # All seven survive, the suggestion is folded in, and the new one is offered.
    assert len(cols) == 8
    assert "爐號" in cols["供應商批號"]["aliases"]
    assert "L/C NO." in cols["供應商批號"]["aliases"]
    assert "熱處理" in cols
    assert r.json()["notes"] == "第 2 份對不起來"


def test_a_reviewed_report_can_be_kept_as_a_sample(client):
    """The reviewer's corrections are the same evidence, and they used to evaporate."""
    job_id = stage(client, markdown="# 南寶")
    client.put(
        f"/api/erp/jobs/{job_id}/rows",
        json={"rows": [{"supplier_lot": "24102102", "test_item": "固成份", "result": "40.12"}]},
    )
    r = client.post(f"/api/erp/jobs/{job_id}/teach")
    assert r.status_code == 201, r.text
    sample_id = r.json()["job_id"]

    from erp import store

    # Stored under the display names — the shape a customer's workbook arrives
    # in — so the draft prompt only ever sees one kind of answer table.
    assert store.get_expected_rows(sample_id)[0]["供應商批號"] == "24102102"
    assert client.get("/api/erp/jobs").json()["count"] == 1  # the sample is not a report


# ── Safety ───────────────────────────────────────────────────────────────────
@pytest.mark.parametrize("job_id", ["..", "../../etc/passwd", "not-hex", "0" * 32])
def test_malformed_or_unknown_job_ids_are_404_not_a_filesystem_walk(client, job_id):
    assert client.get(f"/api/erp/jobs/{job_id}").status_code == 404


def test_oversized_markdown_is_refused(client):
    from erp import routes

    r = client.post(
        "/api/erp/jobs",
        json={
            "documents": [
                {"filename": "big.pdf", "markdown": "x" * (routes.MAX_MARKDOWN_CHARS + 1)}
            ]
        },
    )
    assert r.status_code == 413


def test_deleted_job_is_gone(client):
    job_id = stage(client)
    assert client.delete(f"/api/erp/jobs/{job_id}").status_code == 200
    assert client.get(f"/api/erp/jobs/{job_id}").status_code == 404
